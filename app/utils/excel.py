from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from io import BytesIO

from app.utils.date_utils import utc_to_local_time

def minutes_to_hm(minutes: int | None) -> str:
    if minutes is None:
        return "—"

    h = minutes // 60
    m = minutes % 60
    return f"{h}h {m}m"

def create_attendance_excel(attendance_by_user: dict):
    wb = Workbook()
    ws = wb.active
    ws.title = "Attendance"

    header_font = Font(bold=True)
    center = Alignment(horizontal="center")

    start_col = 1  # Excel column index

    for user in attendance_by_user.values():
        # Merge user name header
        ws.merge_cells(
            start_row=1,
            start_column=start_col,
            end_row=1,
            end_column=start_col + 4,
        )
        cell = ws.cell(row=1, column=start_col, value=user["name"])
        cell.font = header_font
        cell.alignment = center

        # Column headers
        headers = ["Date", "In", "Out", "TH", "OT"]
        for i, h in enumerate(headers):
            c = ws.cell(row=2, column=start_col + i, value=h)
            c.font = header_font
            c.alignment = center

        # Data rows
        row = 3
        for r in user["records"]:
            ws.cell(row=row, column=start_col, value=r["date"].strftime("%d-%m-%Y"))
            ws.cell(row=row, column=start_col + 1, value=utc_to_local_time(r["in_time"]))
            ws.cell(row=row, column=start_col + 2, value=utc_to_local_time(r["out_time"]))
            ws.cell(row=row, column=start_col + 3, value=minutes_to_hm(r["total_minutes"]))
            ws.cell(row=row, column=start_col + 4, value=minutes_to_hm(r["ot_minutes"]))
            row += 1

        start_col += 6  # 5 cols + 1 gap

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer